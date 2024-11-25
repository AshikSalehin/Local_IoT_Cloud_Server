from flask import Flask, request, jsonify, render_template
import openpyxl
import os
import socket
from datetime import datetime

app = Flask(__name__)

db_file = 'iot_database.xlsx'

if not os.path.exists(db_file):
    workbook = openpyxl.Workbook()
    workbook.save(db_file)

def get_user_sheet(username):
    workbook = openpyxl.load_workbook(db_file)
    if username in workbook.sheetnames:
        return workbook[username]
    else:
        return None

def create_user_sheet(username, sensor_names):
    workbook = openpyxl.load_workbook(db_file)
    if username not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=username)
        sheet.append(['Timestamp'] + sensor_names)
        workbook.save(db_file)

def username_exists(username):
    with open('users.txt', 'r') as f:
        users = f.readlines()
    for user in users:
        if user.split(',')[0] == username:
            return True
    return False


def get_local_ip():
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(('10.254.254.254', 1))
        local_ip = s.getsockname()[0]
    except Exception:
        local_ip = '127.0.0.1'
    finally:
        s.close()
    return local_ip

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        sensor_count = int(request.form['sensor_count'])
        sensor_names = [request.form[f'sensor_{i}'] for i in range(sensor_count)]
        
        if username_exists(username):
            return 'Username already exists, please choose a different username.', 409
        
        create_user_sheet(username, sensor_names)
        
        with open('users.txt', 'a') as f:
            f.write(f'{username},{password}\n')
        
        return 'User registered successfully'
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user_exist = username_exists(username)
        
        if user_exist:
            return render_template('profile.html', username=username)
        else:
            return "Error fetching data: " + data.get_data(as_text=True)
    
    return render_template('login.html')

@app.route('/data/<username>', methods=['POST'])
def receive_data(username):
    sheet = get_user_sheet(username)
    if sheet:
        sensor_data = [request.args.get(sensor) for sensor in sheet[1][1:]]
        sheet.append([request.args.get('timestamp')] + sensor_data)
        workbook = sheet.parent
        workbook.save(db_file)
        return 'Data saved successfully'
    else:
        return 'User not found', 404

@app.route('/datas/<username>', methods=['GET'])
def get_data(username):
    sheet = get_user_sheet(username)
    n = int(request.args.get('n'))
    if sheet:
        data = []
        all_rows = list(sheet.iter_rows(values_only=True))
        data = all_rows[-n:]
        return jsonify(data)
    else:
        return 'User not found', 404
    

@app.route('/push_data/<username>', methods=['GET'])
def push_data(username):

    workbook = openpyxl.load_workbook('iot_database.xlsx')

    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Dynamically extract sensor data from the URL
    sensor_data = {key: value for key, value in request.args.items()}
    
    if not sensor_data:
        return "No sensor data provided.", 400

    if username not in workbook.sheetnames:
        return f"User '{username}' not found.", 404

    sheet = workbook[username]

    next_row = sheet.max_row + 1

    sheet[f'A{next_row}'] = timestamp
    for i, (sensor, value) in enumerate(sensor_data.items(), start=1):
        sheet.cell(row=next_row, column=i+1, value=value)

    workbook.save('iot_database.xlsx')

    return "Data inserted successfully", 200


@app.route('/get_ip', methods=['GET'])
def get_ip():
    local_ip = get_local_ip()
    return jsonify({'ip': local_ip})
    
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
